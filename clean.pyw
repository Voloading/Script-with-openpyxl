import os
import sys
import openpyxl as op
from tkinter import *
from tkinter import messagebox
from tkinter import ttk
from openpyxl.styles import Alignment, Font


root = Tk()
root.title("Приложение для любимой")
root.geometry("400x400+400+200")

label = Label(text="Удаление пустых строк в excel")
label.pack(side=TOP)
label = Label(text="Подготовлены следующие файлы:")
label.pack(side=TOP)

current_dir = os.getcwd()
excels = [files for files in os.listdir(current_dir) if files.endswith('.xlsx')]
total__files = len(excels)
result = []
if total__files == 0:
    messagebox.showinfo(f'Ошибка!!!', 'В папке нет excel файлов')
    root.destroy()
    sys.exit()
counter = 1
for file_name in excels:
    label = Label(text=f'{counter}. {file_name}')
    label.pack(side=TOP, anchor=W)
    counter += 1

def deleting_empty_str(start_file_name):
    wb = op.load_workbook(start_file_name)
    ws = wb.active
    row_count = ws.max_row
    down = []
    for col in ws.iter_rows(min_row=row_count-5, max_row=row_count, min_col=1, max_col=20, values_only=True):
        for elem in col:                                #Сохраняем подвал документа
            if elem != None:
                down.append(elem)
    sum = 0
    fix = 0
    for r in range(row_count - 7, 14, -1):              #Проходим по строкам снизу вверх
        for c in range(3, 21):                          #Проходим по столбцам в строке,
            elem = ws.cell(row=r, column=c)             #оставляем те строки, где есть хотя бы одно значение
            if elem.value != None:                      #Удаляем пустые
                sum += 1
                break
        else:
            fix += 1
            ws.delete_rows(r, amount=1)

    row_count = ws.max_row
    for r in range(15, row_count - 6):                                              #форматирование столбца 1 "Порядковый номер"
        ws.cell(row=r, column= 1).value = r - 14                                    #обьединение ячеек
        for d in [(6,7),(8,9),(10,11),(12,14),(15,16),(17,18)]:
            ws.merge_cells(start_row=r, start_column=d[0], end_row=r, end_column=d[1])

    for d in [(8,9),(10,11),(12,14),(15,16),(17,18)]:
        ws.merge_cells(start_row=row_count - 6, start_column=d[0], end_row=row_count - 6, end_column=d[1])           #форматирование строки "Итого"

    ws.delete_rows(row_count - 5, amount=6)             #редактирование подвала документа
    ws.append([])
    ws.append({2: down[0], 12:down[1]})
    ws.merge_cells(start_row=row_count - 4, start_column=2, end_row=row_count - 4, end_column=5)
    ws.cell(row_count - 4, 2).alignment = Alignment(horizontal='center')
    ws.merge_cells(start_row=row_count - 4, start_column=12, end_row=row_count - 4, end_column=16)
    ws.cell(row_count - 4, 12).alignment = Alignment(horizontal='center')
    ws.cell(row_count - 4, 12).font = Font(size=10)

    ws.append({8: down[2], 12:down[3]})
    ws.append([])
    ws.append({2: down[4], 12:down[5]})
    ws.merge_cells(start_row=row_count - 1, start_column=2, end_row=row_count - 1, end_column=5)
    ws.cell(row_count - 1, 2).alignment = Alignment(horizontal='center')
    ws.merge_cells(start_row=row_count - 1, start_column=12, end_row=row_count - 1, end_column=16)
    ws.cell(row_count - 1, 12).alignment = Alignment(horizontal='center')
    ws.cell(row_count - 1, 12).font = Font(size=10)

    ws.append({8: down[6], 12:down[7]})

    new_dir = 'Готовые MO'                              #Создаем папку и сохраняем документ
    path = os.path.join(current_dir, new_dir)
    wb.save(f'{path}/New{start_file_name}')
    result.append(f'{start_file_name} Удалено - {fix}, Сохранено - {sum}\n')
    print(f'Удалено - {fix}, Сохранено - {sum}')

def clicked():
    for excel in excels:
        deleting_empty_str(excel)
    messagebox.showinfo(f'Готово', f'{"\n".join(result)}')
    root.destroy()

button_start = ttk.Button(text = "Старт", command = clicked)
button_start.pack(side=TOP)


root.mainloop()